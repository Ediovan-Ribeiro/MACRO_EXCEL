from openpyxl import load_workbook
import os
from datetime import datetime
import win32com.client as win32  # Para conversão de Excel para PDF

# Caminho para o arquivo Excel
CAMINHO_ENTRADA = os.path.join('..', 'dados', 'entrada', 'exemplo.xlsm')  # Caminho do arquivo com as abas
CAMINHO_SAIDA = os.path.join('..', 'dados', 'saida')                     # Pasta onde os arquivos serão salvos

# Verificar se o arquivo de entrada existe
if not os.path.exists(CAMINHO_ENTRADA):
    print(f"Erro: Arquivo não encontrado: {CAMINHO_ENTRADA}")
    exit()

# Criar a pasta de saída, se não existir
os.makedirs(CAMINHO_SAIDA, exist_ok=True)

# Carregar o arquivo Excel
try:
    wb = load_workbook(CAMINHO_ENTRADA, keep_vba=True)  # keep_vba=True para arquivos .xlsm
    print("Arquivo carregado com sucesso.")
except Exception as e:
    print(f"Erro ao carregar o arquivo: {e}")
    exit()

# Acessar as abas 1OF e MODELO
try:
    aba_1of = wb['1OF']
    aba_modelo = wb['MODELO']
    print("Abas '1OF' e 'MODELO' carregadas com sucesso.")
except KeyError as e:
    print(f"Erro: Aba não encontrada no arquivo: {e}")
    exit()

# Função para transferir dados da aba 1OF para a aba MODELO
def transferir_dados(linha_1of, nova_aba_modelo):
    try:
        # Pegar os valores das células na linha da aba 1OF
        valor_a = aba_1of.cell(row=linha_1of, column=1).value  # Coluna A   tipo de matrícula
        valor_b = aba_1of.cell(row=linha_1of, column=2).value  # Coluna B   número da matrícula
        valor_c = aba_1of.cell(row=linha_1of, column=3).value  # Coluna C   DATA
        valor_d = aba_1of.cell(row=linha_1of, column=4).value  # Coluna D   ÁREA
        valor_e = aba_1of.cell(row=linha_1of, column=5).value  # Coluna E   MEDIDA OFICIAL
        valor_f = aba_1of.cell(row=linha_1of, column=6).value  # Coluna F   CORDENADAS GEODÉSICAS 
        valor_g = aba_1of.cell(row=linha_1of, column=7).value  # Coluna G   INSCRIÇÃO MUNICIPAL
        valor_h = aba_1of.cell(row=linha_1of, column=8).value  # Coluna H   SNCR
        valor_i = aba_1of.cell(row=linha_1of, column=9).value  # Coluna I   NIRF
        valor_j = aba_1of.cell(row=linha_1of, column=10).value  # Coluna J  CAR
        valor_k = aba_1of.cell(row=linha_1of, column=11).value  # Coluna K  SIGEF 
        valor_l = aba_1of.cell(row=linha_1of, column=12).value  # Coluna L  MUNICIPIO / UF
        valor_m = aba_1of.cell(row=linha_1of, column=13).value  # Coluna M  CNS DO REG. ANTERIOR
        valor_n = aba_1of.cell(row=linha_1of, column=14).value  # Coluna N  TIPO DO REG. ANTERIOR
        valor_o = aba_1of.cell(row=linha_1of, column=15).value  # Coluna O  N° DE ORDEM DO REG. ANTERIOR
        valor_p = aba_1of.cell(row=linha_1of, column=16).value  # Coluna P  DATA DO REG. ANTERIOR
        valor_q = aba_1of.cell(row=linha_1of, column=17).value  # Coluna Q  TIPO IMÓVEL
        valor_r = aba_1of.cell(row=linha_1of, column=18).value  # Coluna R  ESPÉCIE DE DOMÍNIO
        valor_s = aba_1of.cell(row=linha_1of, column=19).value  # Coluna S  TIPO DE IMÓVEL PUB.
        valor_t = aba_1of.cell(row=linha_1of, column=20).value  # Coluna T  LEGISLAÇÃO OU ATO ADMINISTRATIVO
        valor_u = aba_1of.cell(row=linha_1of, column=21).value  # Coluna U  ENCERRAMENTO DE REGISTRO
        valor_v = aba_1of.cell(row=linha_1of, column=22).value  # Coluna V  CIRCUNSCRIÇÃO TERRITORIAL
        valor_w = aba_1of.cell(row=linha_1of, column=23).value  # Coluna W  SOBREPOSIÇÃO DE ÁREA
        valor_x = aba_1of.cell(row=linha_1of, column=24).value  # Coluna X  IMÓVEIS EM SOBREPOSIÇÃO
        valor_y = aba_1of.cell(row=linha_1of, column=25).value  # Coluna Y  DUPLICIDADE MATERIAL
        valor_z = aba_1of.cell(row=linha_1of, column=26).value  # Coluna Z  IMÓVEIS EM DUPLICIDADE
        valor_aa = aba_1of.cell(row=linha_1of, column=27).value  # Coluna AA    BLOQUEIO DE REGISTRO

        # Colocar esses valores na aba MODELO nas células correspondentes
        nova_aba_modelo.cell(row=10, column=1, value=valor_a)  # Linha 10, 
        nova_aba_modelo.cell(row=10, column=2, value=valor_b)  # Linha 10, 
        nova_aba_modelo.cell(row=10, column=3, value=valor_c)  # Linha 10, 
        nova_aba_modelo.cell(row=13, column=1, value=valor_d)  # Linha 13, 
        nova_aba_modelo.cell(row=13, column=2, value=valor_e)  # Linha 13, 
        nova_aba_modelo.cell(row=13, column=3, value=valor_f)  # Linha 13, 
        nova_aba_modelo.cell(row=16, column=1, value=valor_g)  # Linha 16, 
        nova_aba_modelo.cell(row=16, column=2, value=valor_h)  # Linha 16, 
        nova_aba_modelo.cell(row=18, column=1, value=valor_j)  # Linha 18, 
        nova_aba_modelo.cell(row=21, column=1, value=valor_k)  # Linha 21, 
        nova_aba_modelo.cell(row=23, column=1, value=valor_i)  # Linha 23, 
        nova_aba_modelo.cell(row=23, column=3, value=valor_l)  # Linha 23, 
        nova_aba_modelo.cell(row=25, column=1, value=valor_o)  # Linha 25, 
        nova_aba_modelo.cell(row=25, column=2, value=valor_p)  # Linha 25, 
        nova_aba_modelo.cell(row=25, column=3, value=valor_n)  # Linha 25, 
        nova_aba_modelo.cell(row=28, column=1, value=valor_m)  # Linha 28, 
        nova_aba_modelo.cell(row=28, column=2, value=valor_q)  # Linha 28, 
        nova_aba_modelo.cell(row=28, column=3, value=valor_r)  # Linha 28, 
        nova_aba_modelo.cell(row=31, column=1, value=valor_s)  # Linha 31, 
        nova_aba_modelo.cell(row=31, column=2, value=valor_t)  # Linha 31, 
        nova_aba_modelo.cell(row=31, column=3, value=valor_u)  # Linha 31, 
        nova_aba_modelo.cell(row=34, column=1, value=valor_v)  # Linha 34, 
        nova_aba_modelo.cell(row=34, column=2, value=valor_w)  # Linha 34, 
        nova_aba_modelo.cell(row=34, column=3, value=valor_x)  # Linha 34, 
        nova_aba_modelo.cell(row=37, column=1, value=valor_y)  # Linha 37, 
        nova_aba_modelo.cell(row=37, column=2, value=valor_z)  # Linha 37, 
        nova_aba_modelo.cell(row=37, column=3, value=valor_aa)  # Linha 37, 
    except Exception as e:
        print(f"Erro ao transferir dados da linha {linha_1of}: {e}")

# Função para salvar apenas a aba MODELO em um novo arquivo Excel
def salvar_apenas_modelo(novo_wb, caminho_arquivo_excel):
    try:
        # Remover todas as abas, exceto a MODELO
        for sheet in novo_wb.sheetnames:
            if sheet != 'MODELO':
                novo_wb.remove(novo_wb[sheet])
        # Salvar o arquivo Excel
        novo_wb.save(caminho_arquivo_excel)
        print(f"Arquivo Excel salvo: {caminho_arquivo_excel}")
    except Exception as e:
        print(f"Erro ao salvar o arquivo Excel: {e}")

# Função para converter Excel para PDF
def excel_para_pdf(caminho_excel, caminho_pdf):
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        workbook = excel.Workbooks.Open(caminho_excel)
        workbook.ExportAsFixedFormat(0, caminho_pdf)  # 0 = Tipo PDF
        workbook.Close()
        excel.Quit()
        print(f"Arquivo PDF salvo: {caminho_pdf}")
    except Exception as e:
        print(f"Erro ao converter Excel para PDF: {e}")

# Definir o intervalo de linhas que você quer processar
total_linhas = aba_1of.max_row

# Processar as linhas da aba 1OF e transferir para a aba MODELO
for linha in range(2, total_linhas + 1):  # Começar na linha 2 (ignorando o cabeçalho)
    # Verificar se a linha está vazia
    if all(cell.value is None for cell in aba_1of[linha]):
        print(f"Linha {linha} está vazia. Ignorando...")
        continue

    # Criar uma cópia do arquivo original
    novo_wb = load_workbook(CAMINHO_ENTRADA, keep_vba=True)
    nova_aba_1of = novo_wb['1OF']
    nova_aba_modelo = novo_wb['MODELO']

    # Transferir dados para a aba MODELO
    transferir_dados(linha, nova_aba_modelo)

    # Salvar apenas a aba MODELO em um novo arquivo Excel
    nome_arquivo_excel = f"modelo_linha_{linha}.xlsm"
    caminho_arquivo_excel = os.path.join(CAMINHO_SAIDA, nome_arquivo_excel)
    salvar_apenas_modelo(novo_wb, caminho_arquivo_excel)

    # Converter o arquivo Excel para PDF
    nome_arquivo_pdf = f"modelo_linha_{linha}.pdf"
    caminho_arquivo_pdf = os.path.join(CAMINHO_SAIDA, nome_arquivo_pdf)
    excel_para_pdf(caminho_arquivo_excel, caminho_arquivo_pdf)

    print(f"Arquivos salvos: {caminho_arquivo_excel} e {caminho_arquivo_pdf}")

print("Processo concluído! Todos os arquivos foram salvos.")